from django.views.generic import TemplateView
from web_project import TemplateLayout
from django.db.models import Q
from .models import Candidato, Vaga
from .forms import CandidatoForm, VagaForm
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import View
import os
from django.conf import settings
from django.core.files.base import ContentFile
from docx import Document
import pdfkit
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib.auth import login
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.contrib.auth.forms import AuthenticationForm
from django.db.models import Count
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout


def home_view(request):
    if not request.user.is_authenticated:
        return redirect('login')
    return redirect('index')


@require_http_methods(["GET", "POST"])
def login_view(request):
    context = {'layout_path': 'base.html',}
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            print(user)
            login(request, user)
            return redirect('index')
        else:
            print(form.errors)
            messages.error(request, "Usuário ou senha inválidos.")
    else:
        form = AuthenticationForm()

    return render(request, 'login.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('login')


def export_to_xlsx(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Currículos e Vagas"

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def style_cell(cell, bold=False, alignment=None, border=None):
        if bold:
            cell.font = bold_font
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border

    ws.append(['Currículos'])
    ws.merge_cells('A1:B1')
    style_cell(ws['A1'], bold=True, alignment=center_alignment)

    ws.append(['Categoria', 'Quantidade'])
    for cell in ws[2]:
        style_cell(cell, bold=True, alignment=center_alignment, border=thin_border)

    ws.append(['Aprovados', Candidato.objects.filter(resultado='aprovado').count()])
    ws.append(['Reprovados', Candidato.objects.filter(resultado='reprovado').count()])
    ws.append(['Vaga Futura', Candidato.objects.filter(resultado='futura').count()])
    ws.append(['Entrevistados', Candidato.objects.filter(processo_selecao='entrevistado').count()])
    ws.append(['Aguardando Entrevista', Candidato.objects.filter(processo_selecao='aguardando_entrevista').count()])
    ws.append(['Total de Currículos', Candidato.objects.count()])

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            style_cell(cell, alignment=center_alignment, border=thin_border)

    ws.append([])

    start_row = ws.max_row + 1
    ws.append(['Vagas'])
    ws.merge_cells(f'A{start_row}:B{start_row}')
    style_cell(ws[f'A{start_row}'], bold=True, alignment=center_alignment)

    ws.append(['Categoria', 'Quantidade'])
    for cell in ws[start_row + 1]:
        style_cell(cell, bold=True, alignment=center_alignment, border=thin_border)

    ws.append(['Vagas Abertas', Vaga.objects.filter(status='aberta').count()])
    ws.append(['Vagas Preenchidas', Vaga.objects.filter(status='fechada').count()])
    ws.append(['Total de Vagas', Vaga.objects.count()])

    for row in ws.iter_rows(min_row=start_row + 2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            style_cell(cell, alignment=center_alignment, border=thin_border)

    ws.append([])
    ws.append(['Título da Vaga', 'Quantidade em Aberto'])
    for cell in ws[ws.max_row]:
        style_cell(cell, bold=True, alignment=center_alignment, border=thin_border)

    vagas_acumuladas = Vaga.objects.filter(status='aberta').values('titulo').annotate(quantidade=Count('titulo'))
    for vaga in vagas_acumuladas:
        ws.append([vaga['titulo'], vaga['quantidade']])

    for row in ws.iter_rows(min_row=ws.max_row - len(vagas_acumuladas) + 1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            style_cell(cell, alignment=center_alignment, border=thin_border)

    for column_cells in ws.columns:
        max_length = 0
        column = [cell for cell in column_cells if not isinstance(cell, type(ws['A1']))]
        if not column:
            continue
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=curriculos_e_vagas.xlsx'
    wb.save(response)

    return response


def convert_and_save_docx_as_pdf(candidato, arquivo):
    base_filename = os.path.splitext(arquivo.name)[0]
    output_filename = f"{candidato.nome.replace(' ', '_')}.pdf"
    output_path = os.path.join(settings.MEDIA_ROOT, 'curriculos', output_filename)

    try:

        doc = Document(arquivo)

        html_content = '<html><body>'
        for para in doc.paragraphs:
            html_content += f'<p>{para.text}</p>'
        html_content += '</body></html>'

        temp_html_path = os.path.join(settings.MEDIA_ROOT, 'curriculos', f'{base_filename}.html')
        with open(temp_html_path, 'w') as f:
            f.write(html_content)

        pdfkit.from_file(temp_html_path, output_path)

        with open(output_path, 'rb') as pdf_file:
            candidato.curriculo.save(output_filename, ContentFile(pdf_file.read()))

        print(f"Arquivo salvo como PDF no modelo Candidato com o nome: {output_filename}")
        return output_filename

    except Exception as e:
        print(f"Erro ao converter e salvar o arquivo: {e}")
        return None

    finally:

        if os.path.exists(temp_html_path):
            os.remove(temp_html_path)


@login_required
def index(request):
    if request.method == 'POST':
        form = CandidatoForm(request.POST, request.FILES)
        if form.is_valid():
            candidato = form.save(commit=False)

            if 'curriculo' in request.FILES:
                uploaded_file = request.FILES['curriculo']
                candidato.curriculo.save(uploaded_file.name, uploaded_file)
                print(f"Arquivo salvo em: {candidato.curriculo.path}")

                output_filename = convert_and_save_docx_as_pdf(candidato, uploaded_file)
                if output_filename:
                    candidato.curriculo.name = f'curriculos/{output_filename}'

            candidato.save()
            messages.success(request, "Candidato cadastrado com sucesso!")
            return redirect('index')
        else:
            messages.error(request, "Erro ao cadastrar candidato. Por favor, verifique os dados.")
    else:
        form = CandidatoForm()

    layout_path = 'base.html'
    return render(request, 'index.html', {'form': form, 'layout_path': layout_path})


class SampleView(TemplateView):
    template_name = "index.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context = TemplateLayout.init(self, context)
        context['menu_data'] = {
            "menu": [
                {
                    "url": "index",
                    "icon": "menu-icon tf-icons bx bx-home-circle",
                    "name": "Página Inicial",
                    "slug": "index"
                },
                {
                    "url": "page-2",
                    "icon": "menu-icon tf-icons bx bx-search",
                    "name": "Buscar Currículo",
                    "slug": "page-2"
                },
                {
                    "url": "page-3",
                    "icon": "menu-icon tf-icons bx bx-bar-chart-alt-2",
                    "name": "Acompanhamento RH",
                    "slug": "page-3"
                },
                {
                    "url": "manage-jobs",
                    "icon": "menu-icon tf-icons bx bx-briefcase",
                    "name": "Gerenciar Vagas",
                    "slug": "manage-jobs"
                }
            ]
        }

        return context

    def post(self, request, *args, **kwargs):
        form = CandidatoForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Candidato cadastrado com sucesso!')
                return redirect('index')
            except Exception as e:
                messages.error(request, f'Ocorreu um erro ao cadastrar o candidato: {e}')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')

        context = self.get_context_data(**kwargs)
        context['form'] = form
        return self.render_to_response(context)


class BuscarCurriculoView(TemplateView):
    template_name = 'page_2.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context = TemplateLayout.init(self, context)
        filtros = Q()

        if self.request.GET.get('nome'):
            filtros &= Q(nome__icontains=self.request.GET['nome'])
        if self.request.GET.get('email'):
            filtros &= Q(email__icontains=self.request.GET['email'])
        if self.request.GET.get('telefone'):
            filtros &= Q(telefone__icontains=self.request.GET['telefone'])
        if self.request.GET.get('vaga_interesse'):
            filtros &= Q(vaga_interesse__icontains=self.request.GET['vaga_interesse'])

        context['resultados'] = Candidato.objects.filter(filtros)

        return context


class DashboardRHView(TemplateView):
    template_name = 'page_3.html'

    def get_context_data(self, **kwargs):
        context = TemplateLayout.init(self, super().get_context_data(**kwargs))

        context['total_cadastrados'] = Candidato.objects.count()
        context['total_aprovados'] = Candidato.objects.filter(resultado='aprovado').count()
        context['total_reprovados'] = Candidato.objects.filter(resultado='reprovado').count()
        context['total_vaga_futura'] = Candidato.objects.filter(resultado='futura').count()
        context['total_entrevistados'] = Candidato.objects.filter(processo_selecao='entrevistado').count()
        context['total_aguardando_entrevista'] = Candidato.objects.filter(processo_selecao='aguardando_entrevista').count()

        return context


class ManageJobsView(TemplateView):
    template_name = 'manage_jobs.html'

    def get_context_data(self, **kwargs):
        context = TemplateLayout.init(self, super().get_context_data(**kwargs))
        context['vagas'] = Vaga.objects.all()
        context['form'] = VagaForm()

        context['total_vagas'] = Vaga.objects.count()
        context['vagas_abertas'] = Vaga.objects.filter(status='aberta').count()
        context['vagas_fechadas'] = Vaga.objects.filter(status='fechada').count()

        return context

    def post(self, request, *args, **kwargs):
        if 'fechar_vaga' in request.POST:
            vaga_id = request.POST.get('vaga_id')
            vaga = get_object_or_404(Vaga, id=vaga_id)
            vaga.status = 'fechada'
            vaga.save()
            return redirect('manage-jobs')

        form = VagaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('manage-jobs')
        else:
            messages.error(request, 'Erro ao cadastrar a vaga. Verifique os dados.')
        return self.get(request, *args, **kwargs)


class DeleteVagaView(View):
    def post(self, request, *args, **kwargs):
        vaga = get_object_or_404(Vaga, pk=kwargs['id'])
        vaga.delete()
        messages.success(request, 'Vaga excluída com sucesso!')
        return redirect('manage-jobs')


class EditVagaView(TemplateView):
    template_name = 'manage_jobs.html'

    def post(self, request, *args, **kwargs):
        vaga = get_object_or_404(Vaga, id=kwargs['id'])
        form = VagaForm(request.POST, instance=vaga)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vaga atualizada com sucesso!')
        else:
            messages.error(request, 'Erro ao atualizar a vaga. Verifique os dados.')

        context = self.get_context_data(**kwargs)
        context['form'] = form
        context['vagas'] = Vaga.objects.all()
        return redirect('manage-jobs')


class EditCandidatoView(TemplateView):
    template_name = 'page_2.html'

    def post(self, request, *args, **kwargs):
        candidato = get_object_or_404(Candidato, id=kwargs['id'])
        form = CandidatoForm(request.POST, request.FILES, instance=candidato)
        if form.is_valid():
            form.save()
            messages.success(request, 'Candidato atualizado com sucesso!')
        else:
            messages.error(request, 'Erro ao atualizar candidato. Verifique os dados.')

        context = self.get_context_data(**kwargs)
        context['form'] = CandidatoForm()
        context['resultados'] = Candidato.objects.all()
        return redirect('page-2')


class DeleteCandidatoView(View):
    def post(self, request, *args, **kwargs):
        candidato = get_object_or_404(Candidato, pk=kwargs['id'])
        candidato.delete()
        messages.success(request, 'Candidato excluído com sucesso!')
        return redirect('page-2')
